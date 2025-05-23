#!/usr/bin/env python3
import os
import dropbox
from dotenv import load_dotenv
import mysql.connector
from openpyxl import load_workbook
from datetime import datetime
from dropbox.files import WriteMode
from secrets_loader import get_secret


#pip install mysql-connector-python openpyxl
#save your excel template in project root folder

#1.connect to MySQL
DB_CONFIG = {
    'host':     '127.0.0.1',         #local machine's mysql
    'port':     3007,                #local machine's port      
    'user':     'wpuser',
    'password': 'wppassword',
    'database': 'wpdb',
}

#2.Excel template path and output 
TEMPLATE_PATH = 'Report_Template_without_pics.xlsx'
#TO DO: output file format= machineName_id_datetime_
OUTPUT_PATH = f'filled_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

#load .env
load_dotenv()

# ACCESS_TOKEN = os.getenv('DROPBOX_TOKEN')
DROPBOX_FOLDER = os.getenv('DROPBOX_FOLDER', '/')

if not ACCESS_TOKEN:
    raise RuntimeError("请在 .env 文件中设置 DROPBOX_TOKEN")

def fetch_latest_submission(submission_id=None):
    """Fetch one submission record from the database."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    sql = "SELECT id, user_id, field_1, field_2, created_at FROM wp_submissions"
    params = ()
    if submission_id:
        sql += " WHERE id = %s"
        params = (submission_id, )
    sql += " ORDER BY created_at DESC LIMIT 1"
    cursor.execute(sql, params)
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    return row

# function: fill value into Excel
def fill_excel(row):
    """load the template, write data into cells, and save a new file"""
    if not row:
        raise ValueError("no data found in the database.")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws['H22'] = row['field_1']
    ws['A34'] = row['field_2']

    wb.save(OUTPUT_PATH)
    print(f'generated excel file: {OUTPUT_PATH}')

def upload_file_to_dropbox(local_path: str) -> str:
    """
    将本地文件上传到 DROPBOX_FOLDER，然后返回在 Dropbox 上的完整路径。
    使用 overwrite 模式：如果远端已有同名文件，将被覆盖。
    """
    if not os.path.isfile(local_path):
        raise FileNotFoundError(f"本地文件不存在：{local_path}")

    dbx = dropbox.Dropbox(ACCESS_TOKEN)
    filename = os.path.basename(local_path)
    dropbox_path = os.path.join(DROPBOX_FOLDER, filename).replace("\\", "/")

    # 以覆盖模式上传
    with open(local_path, 'rb') as f:
        dbx.files_upload(f.read(), dropbox_path, mode=WriteMode('overwrite'))

    return dropbox_path

# main: run the script
if __name__== "__main__":
    #Fetch secret from AWS
    try:
        my_secret = get_secret()
        print("Secret retrieved successfully.")  # optional, for debug only
    except Exception as e:
        print(f"Failed to retrieve secret: {e}")
        exit(1)  # Exit if secret is critical

    data = fetch_latest_submission()
    fill_excel(data)
    try:
        dest = upload_file_to_dropbox(OUTPUT_PATH)
        print(f'已上传到 Dropbox: {dest}')
    except Exception as e:
        print(f"upload failure: {e}")