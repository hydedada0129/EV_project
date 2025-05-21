#!/usr/bin/env python3
import os
import mysql.connector
import datetime
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from upload_to_db import upload_file_to_dropbox

# connect to MySQL
DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 3007,
    'user': 'wpuser',
    'password': 'wppassword',
    'database': 'wpdb',
}

TEMPLATE_PATH = 'Report_Template_without_pics.xlsx'
PHOTO_POSITIONS = ["M5", "S5", "Y5", "M26", "S26", "Y26", "M47", "S47", "Y47"]
# PHOTO_POSITIONS = ["M5", "M26", "M47", "M68"]


def fetch_latest_submission():
    """直接從 wp_submissions 表格獲取最新提交的資料"""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    
    # 只查詢 wp_submissions 表格，查詢過去 5 分鐘內提交的資料
    # 註：此處 5 分鐘僅為測試，實際使用時會改為 1 天
    sql = """
        SELECT id, user_id, machine_name, job_number, date_submitted, site_address,
               model_number, serial_number, equipment_type,
               travel_date, travel_hours, travel_arrived, travel_miles,
               onsite_date, onsite_start, onsite_end, work_description, 
               created_at, photo_url
        FROM wp_submissions
        WHERE created_at >= NOW() - INTERVAL 60 MINUTE
    """
    
    try:
        cursor.execute(sql)
        rows = cursor.fetchall()
        return rows
    finally:
        cursor.close()
        conn.close()


def resize_image_to_fit(img_path, max_width=320, max_height=400):
    """縮放圖片以適合 LibreOffice Calc，保留原始格式"""
    try:
        with PILImage.open(img_path) as img:
            # 保存原始資訊用於除錯
            original_size = img.size
            original_format = img.format
            original_mode = img.mode
            print(f"原始圖片資訊: 尺寸={original_size}, 格式={original_format}, 模式={original_mode}")
            
            # 縮放圖片
            img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
            
            # 生成臨時檔案名，保留原始副檔名
            file_ext = os.path.splitext(img_path)[1]
            if not file_ext:
                # 如果沒有副檔名，根據格式設置
                file_ext = '.png' if original_format == 'PNG' else '.jpg'
                
            resized_path = f"temp_resized_{os.path.basename(img_path)}{file_ext}"
            
            # 處理透明通道但不轉換格式
            if original_format == 'PNG' and (img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info)):
                # 對於 PNG 格式，直接保存，保留透明度
                img.save(resized_path, 'PNG')
                print(f"圖片已縮放，保存為原始 PNG 格式: {resized_path}")
            else:
                # 其他格式也保持不變
                img.save(resized_path, original_format if original_format else 'JPEG')
                print(f"圖片已縮放，保存為原始格式 {original_format}: {resized_path}")
            
            return resized_path
    except Exception as e:
        print(f"圖片處理失敗: {e}")
        import traceback
        traceback.print_exc()
        return None

def download_images(photo_urls):
    """下載圖片到臨時資料夾"""
    if not photo_urls:
        print("沒有照片需要下載")
        return []
        
    # 確保 photo_urls 是字串，如果是則分割成列表
    if isinstance(photo_urls, str):
        photo_urls = photo_urls.split(',')
        print(f"分割 URL 字串，找到 {len(photo_urls)} 個 URL")
    
    # 清理所有空白 URL
    photo_urls = [url.strip() for url in photo_urls if url.strip()]
    print(f"清理後剩餘 {len(photo_urls)} 個有效 URL")
    
    # 創建臨時目錄
    os.makedirs("temp_photos", exist_ok=True)
    local_paths = []
    
    # 實際下載圖片
    for idx, url in enumerate(photo_urls):
        try:
            import requests
            print(f"處理圖片 {idx+1}/{len(photo_urls)}: {url}")
            
            # 修正 localhost URL
            if 'localhost' in url:
                # 從 http://localhost:8080/wp-content/... 轉換為 /home/oem/wordpress-docker/wp-data/wp-content/...
                local_path = url.replace('http://localhost:8080/wp-content', '/home/oem/wordpress-docker/wp-data/wp-content')
                print(f"檢查本地檔案路徑: {local_path}")
                if os.path.exists(local_path):
                    # 直接複製檔案而不是下載
                    import shutil
                    img_path = f"temp_photos/image_{idx}{os.path.splitext(local_path)[1]}"
                    shutil.copy(local_path, img_path)
                    local_paths.append(img_path)
                    print(f"複製本地檔案成功: {img_path}")
                    continue
                else:
                    print(f"本地檔案不存在，嘗試透過 HTTP 下載")
            
            # 正常 HTTP 下載流程
            response = requests.get(url)
            if response.status_code == 200:
                # 根據URL確定檔案副檔名
                ext = os.path.splitext(url)[1]
                if not ext:
                    ext = '.jpg'  # 預設副檔名
                img_path = f"temp_photos/image_{idx}{ext}"
                
                with open(img_path, 'wb') as f:
                    f.write(response.content)
                local_paths.append(img_path)
                print(f"圖片下載成功: {img_path}")
            else:
                print(f"圖片下載失敗，狀態碼: {response.status_code}")
        except Exception as e:
            print(f"圖片下載失敗: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"共下載 {len(local_paths)} 張圖片")
    return local_paths


def cleanup_temp_photos():
    """清理臨時照片資料夾"""
    import shutil
    if os.path.exists("temp_photos"):
        shutil.rmtree("temp_photos")
        print("已刪除 temp_photos 資料夾。")
    
    # 清理暫存的縮放圖片
    for file in os.listdir('.'):
        if file.startswith('temp_resized_'):
            try:
                os.remove(file)
                print(f"已刪除暫存圖片: {file}")
            except Exception as e:
                print(f"刪除暫存圖片失敗: {e}")


def fill_excel(row, output_path):
    """根據提交的資料填充 Excel 模板"""
    if not row:
        raise ValueError("No data found in the database.")

    print(f"處理資料: {row['machine_name']}, ID: {row['id']}")
    
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = row['machine_name']

    # 填充資料到對應的單元格
    ws['G3'] = row['job_number']
    ws['G5'] = row['date_submitted']
    ws['G7'] = row['site_address']
    ws['F22'] = row['model_number']
    ws['H22'] = row['serial_number']
    ws['J22'] = row['equipment_type']
    ws['C27'] = row['travel_date']
    ws['D27'] = row['travel_hours']
    ws['E27'] = row['travel_arrived']
    ws['F27'] = row['travel_miles']
    ws['H27'] = row['onsite_date']
    ws['I27'] = row['onsite_start']
    ws['J27'] = row['onsite_end']
    ws['A34'] = row['work_description']

    # 插入圖片並固定欄寬（防止被圖片自動撐寬）
    fixed_columns = ["M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA"]
    for col in fixed_columns:
        ws.column_dimensions[col].width = 15

    # 處理照片
    if row.get('photo_url'):
        print(f"發現照片: {row['photo_url']}")
        image_paths = download_images(row['photo_url'])
        print(f"下載的圖片數量: {len(image_paths)}")
        
        # 計算可用的照片位置數量
        available_positions = len(PHOTO_POSITIONS)
        print(f"可用的照片位置數量: {available_positions}")
        
        # 計算要處理的照片數量
        photos_to_process = min(len(image_paths), available_positions)
        print(f"將處理 {photos_to_process} 張照片")
        
        # 逐一將照片插入到 Excel
        processed_count = 0
        for i in range(photos_to_process):
            try:
                img_path = image_paths[i]
                print(f"處理第 {i+1} 張照片，原始路徑: {img_path}")
                
                # 縮放圖片並轉換格式
                resized_path = resize_image_to_fit(img_path)
                if not resized_path:
                    print(f"圖片處理失敗，跳過此圖片")
                    continue
                
                print(f"縮放後的圖片路徑: {resized_path}")
                
                # 獲取下一個可用的位置
                position = PHOTO_POSITIONS[processed_count]
                
                # 插入圖片到 Excel
                img = XLImage(resized_path)
                ws.add_image(img, position)
                print(f"已將圖片添加到位置 {position}")
                
                # 增加已處理的計數
                processed_count += 1
                
            except Exception as e:
                print(f"插入第 {i+1} 張照片時發生錯誤: {e}")
                import traceback
                traceback.print_exc()
        
        print(f"成功插入 {processed_count} 張照片到 Excel")
    else:
        print("沒有找到照片")

    wb.save(output_path)
    print(f'已生成 Excel 檔案: {output_path}')
    return output_path


def main():
    """主函數，負責整個處理流程"""
    try:
        print("開始獲取最新提交的資料...")
        data = fetch_latest_submission()
        
        if not data:
            print("未找到任何新提交的資料，程序退出。")
            return
            
        print(f"找到 {len(data)} 筆資料")
        
        for row in data:
            # 格式化時間戳記
            if isinstance(row['created_at'], str):
                created_at = datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S')
            else:
                created_at = row['created_at']
                
            created_at_str = created_at.strftime('%Y%m%d_%H%M%S')
            output_path = f"{row['machine_name']}_{created_at_str}.xlsx"
            
            print(f"處理報告: {output_path}")
            filled_file = fill_excel(row, output_path)
            
            try:
                print(f"上傳 Excel 報告到 Dropbox...")
                dest = upload_file_to_dropbox(output_path)
                print(f'Excel 報告已上傳: {dest}')
            except Exception as e:
                print(f"上傳到 Dropbox 失敗: {e}")
                print(f"Excel 報告已生成，但未上傳: {output_path}")

        cleanup_temp_photos()
        print("處理完成")

    except Exception as e:
        print(f"執行失敗：{e}")
        import traceback
        traceback.print_exc()
        
        # 即使在出現異常時也清理臨時檔案
        try:
            print("嘗試清理臨時檔案...")
            cleanup_temp_photos()
            print("臨時檔案清理完成。")
        except Exception as cleanup_error:
            print(f"清理臨時檔案時發生錯誤: {cleanup_error}")
        
        exit(1)


# 設置信號處理器，以處理意外終止程序的情況（如按 Ctrl+C）
def setup_signal_handlers():
    """設置信號處理器，確保程序被中斷時也能清理臨時檔案"""
    import signal
    import sys
    
    def signal_handler(sig, frame):
        print("\n程序被中斷，正在清理臨時檔案...")
        try:
            cleanup_temp_photos()
            print("臨時檔案清理完成。")
        except Exception as e:
            print(f"清理臨時檔案時發生錯誤: {e}")
        sys.exit(0)
    
    # 註冊 SIGINT (Ctrl+C) 和 SIGTERM 信號處理器
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)


if __name__ == "__main__":
    # 設置信號處理器
    setup_signal_handlers()
    
    # 運行主程序
    main()